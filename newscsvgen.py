#News CSV generator for highlights, loops through all entries on the news items list supplied by Martin/Connell and creates a CSV file for each PA in their respective folder. PSLs will then add info to these spreadsheets so that the newsxmlgen script can be instructed correctly
#Developed by Daniel Hutchings

import csv
import pandas as pd
from pandas.tseries.offsets import BMonthEnd
import glob
import os
import fnmatch
import re
import time
import datetime
import sys
import xml.etree.ElementTree as ET
from lxml import etree
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font, colors


def xls_add_format(PA):
    constantPA = PA  
    XLSFilepath = outputDir + constantPA + '\\' + constantPA + ' news items ' + highlightDate + '.xlsx'

    wb = openpyxl.load_workbook(filename = XLSFilepath)        
    worksheet = wb.active
    worksheet.column_dimensions['F'].hidden= True #hide column
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    #for col in worksheet.columns:
    #    max_length = 0
    #    column = str(col[0].column) # Get the column name
    #    print('Column: ' + column)
    #    if column in ['B', 'F', 'G']: worksheet.column_dimensions[column].width = 50
    #    else:
    #        if column == 'C': worksheet.column_dimensions[column].width = 60
            #else: worksheet.column_dimensions[column].width = 20

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 60
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 50
    worksheet.column_dimensions['G'].width = 50
    worksheet.column_dimensions['H'].width = 20
    worksheet.column_dimensions['I'].width = 20
    worksheet.column_dimensions['J'].width = 20
        
    worksheet.sheet_view.zoomScale = 80
    #print(column)
    ListOfRowsWithDateDiffs = []
    for rows in worksheet.iter_rows(min_row=1, min_col=1):
        for cell in rows:                
            rowNumber = re.search('^\D(.*)',cell.coordinate).group(1)
            colNumber = re.search('(^\D).*',cell.coordinate).group(1)
            
            worksheet[cell.coordinate].alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            worksheet[cell.coordinate].border = thin_border
            if (int(rowNumber) % 2) == '0': #check if row is even number
                worksheet[cell.coordinate].fill = PatternFill(fgColor="ccddff", fill_type = "solid") #if yes fill cell contents with color
            else:
                worksheet[cell.coordinate].fill = PatternFill(fgColor="ffffff", fill_type = "solid") #turn white background
            if colNumber == 'K':
                if str(cell.value) == 'True':                    
                    ListOfRowsWithDateDiffs.append(rowNumber)
                    print(str(cell.value))
                    worksheet[cell.coordinate].fill = PatternFill(fgColor="ffff00", fill_type = "solid") #if yes fill cell contents with color

            if colNumber == 'J': 
                #print('cell %s %s' % (cell.coordinate,cell.value))    
                worksheet[cell.coordinate].hyperlink = cell.value
                worksheet[cell.coordinate].value="View on PSL"   
                worksheet[cell.coordinate].font = Font(color=colors.BLUE, bold=True) 
                #print('cell %s %s' % (cell.coordinate,cell.value))    
                #print(colNumber)  
                
    #loop through rows again to highlight any rows that have dates that are different from eachother
    for rows in worksheet.iter_rows(min_row=1, min_col=1):
        for cell in rows:                
            rowNumber = re.search('^\D(.*)',cell.coordinate).group(1)
            if rowNumber in ListOfRowsWithDateDiffs:
                worksheet[cell.coordinate].fill = PatternFill(fgColor="ffff00", fill_type = "solid") 

    #worksheet.set_column(excel_header, 20)

    worksheet.column_dimensions['K'].hidden= True #hide column    worksheet.auto_filter.ref = "A:J"
    try: wb.save(XLSFilepath)
    except PermissionError: 
        log('COULD NOT EXPORT DUE TO PERMISSION ERROR: ' + XLSFilepath)
    #wait = input("PAUSED...when ready press enter")

def csv_generation(PA, highlightDate, highlightType, df, outputDir, PrevHighlightsFilepath):
    #log('Loading xml: ' + PA + ' ' + PrevHighlightsFilepath)
    #if PrevHighlightsFilepath != 'na':
    #    tree = etree.parse(PrevHighlightsFilepath)
    #    root = tree.getroot()
    #    prev_week_doc_text = str(etree.tostring(root))
        


    constantPA = PA            
    #CSVFilepath = outputDir + constantPA + '\\' + constantPA + ' news items ' + highlightDate + '.csv'
    XLSFilepath = outputDir + constantPA + '\\' + constantPA + ' news items ' + highlightDate + '.xlsx'    
    dfPA = pd.DataFrame()

    if PA == 'Life Sciences and Pharmaceuticals': PA = 'Life Sciences'
    if PA == 'Banking and Finance': PA = 'Banking & Finance'
    if PA == 'Share Schemes': PA = 'Share Incentives'
    if PA == 'Risk and Compliance': PA = 'Risk & Compliance'
    if PA == 'Restructuring and Insolvency': PA = 'Restructuring & Insolvency'
    if PA == 'Wills and Probate': PA = 'Wills & Probate'
    
    if PA == 'Insurance': 
        dfPA = df[df.PA.isin(['Insurance', 'Insurance & Reinsurance', 'Insurance &amp; Reinsurance'])]
    else: 
        if PA == 'Personal Injury': dfPA = df[df.PA.isin(['PI &amp; Clinical Negligence', 'PI & Clinical Negligence', 'Personal Injury'])]
        else: dfPA = df[(df.PA ==PA)] 
    
    #loop through the filtered by PA dataframe checking each citation to see whether it appears in last week's highlights doc text, if it does, don't add it to new dataframe df_temp
    #if PrevHighlightsFilepath != 'na':
    #    log('Checking citations against last week highlight doc...')
    #    df_temp = pd.DataFrame()
    #    dfPA = dfPA.reset_index(drop=True)              #

    #    for index, row in dfPA.iterrows():  
    #        print(dfPA.Citation.iloc[index])        
    #        if prev_week_doc_text.find(str(dfPA.Citation.iloc[index])) > -1:
    #            log('FOUND IN LAST WEEKS ' + PA + ' HIGHLIGHTS: ' + dfPA.Citation.iloc[index])
    #            #log(prev_week_doc_text)
    #            #dfPA.drop(row, inplace=True)
    #            dfPA.loc[index,'Duplicate'] = True
    #            #log('REMOVED FROM DATAFRAME')
    #        else:                
    #            dfPA.loc[index,'Duplicate'] = False
    #        #else:
    #        #    df_temp = df_temp.append(row, ignore_index=True)
    #    #dfPA = df_temp
    #    try:dfPA = dfPA[dfPA.Duplicate==False]
    #    except:log('Error filtering on duplicate = False for ' + PA )
    #else:
    #    log('No highlight doc found for last week, skipping citation overlap check...')

    #dfPA = dfPA.sort_values(['Title'], ascending = True)
    #dfPA.to_csv(CSVFilepath, index=False, encoding='utf-8')

    try:dfPA = dfPA.sort_values(['IssueDate'], ascending = False)
    except: log(str(dfPA.head(5)))
    try: dfPA.to_excel(XLSFilepath, index=False, encoding='utf-8')
    #print('CSV exported to...' + CSVFilepath)
    except PermissionError: 
        log('COULD NOT REFORMAT DUE TO PERMISSION ERROR: ' + XLSFilepath)
    #log('CSV exported to...' + CSVFilepath)
    print('XLS exported to...' + XLSFilepath)
    log('XLS exported to...' + XLSFilepath)
    

def IsWednesday(givendate):
    #print(givendate.weekday())
    if givendate.weekday() == 2: return True
    else: return False
def IsThursday(givendate):
    #print(givendate.weekday())
    if givendate.weekday() == 3: return True
    else: return False


def FindLastWeekHighLightDoc(directory, PA):    
    filelist = glob.iglob(os.path.join(directory, '*.xml')) #builds list of file in a directory based on a pattern
    for filepath in filelist:
        if filepath.find('0S4D.xml') == -1:
            #print(filepath)
            tree = etree.parse(filepath)
            root = tree.getroot()
            metaData = root.xpath("//header:metadata-item[@name='master-topic-link-parameters-3']", namespaces=NSMAP)
            if len(metaData) == 0: metaData = root.xpath("//header:metadata-item[@name='topic-link-parameters-3']", namespaces=NSMAP)

            #print(etree.tostring(metaData[0]))
            try: metaDataValue = re.search('^::([^:]*):',metaData[0].get('value')).group(1)
            except AttributeError: metaDataValue = re.search('^([^:]*):',metaData[0].get('value')).group(1)
            metaDataValue = re.sub(' $', '', metaDataValue) #remove whitespace at the end of the string if present
            if metaDataValue == 'IP and IT': metaDataValue = 'IP'
            if metaDataValue == 'Share Incentives': metaDataValue = 'Share Schemes'
            if metaDataValue == 'InHouse Advisor': metaDataValue = 'In-House Advisor'
            if metaDataValue == 'Life Sciences': metaDataValue = 'Life Sciences and Pharmaceuticals'
            #print(metaDataValue, PA)
            if metaDataValue == PA: 
                #print('Match found returning filepath: ' + filepath)
                return filepath        
    return 'na'
    


def extract_from_news_feed(df_archive):

    df = pd.DataFrame()
    log('Looping through the news xml log...')
    newsListXML = etree.parse(newsListFilepath)
    for newsItem in newsListXML.findall(".//document"):
        #get date and citation info first  
        newsCitation = newsItem.find(".//citation")
        newsCitation = newsCitation.text      
        newsPubDate = newsItem.find(".//publish-date")
        newsPubDate = re.search('([^T]*)T',newsPubDate.text).group(1)
        year = int(re.search('(\d\d\d\d)-(\d\d)-(\d\d)',newsPubDate).group(1))
        month = int(re.search('(\d\d\d\d)-(\d\d)-(\d\d)',newsPubDate).group(2))
        day = int(re.search('(\d\d\d\d)-(\d\d)-(\d\d)',newsPubDate).group(3))
        
        if any(df_archive['Citation'] == newsCitation) == False:
            newsDate = newsItem.find(".//date")
            newsDate = re.search('([^T]*)T',newsDate.text).group(1)
            newsURLString = ''
            newsTitle = newsItem.find(".//title")
            newsTitle = str(newsTitle.text)
            newsLink = search + newsTitle            
            newsMiniSummary = newsItem.find(".//mini-summary")
            newsMiniSummary = str(newsMiniSummary.text)
            newsMiniSummary = re.sub("^.*analysis: ", "", newsMiniSummary)    
            newsMiniSummary = re.sub("^.*Analysis: ", "", newsMiniSummary)                
            if newsDate != newsPubDate: DatesDifferent = "True"
            else: DatesDifferent = "False"
            try: newsSources = etree.tostring(newsItem.find(".//source-links"), encoding="unicode")
            except TypeError: newsSources = ''
            for newsURL in newsItem.findall(".//url"): newsURLString += str(newsURL.get('address')) + ' \n\n'
            for newsPA in newsItem.findall(".//practice-area"):
                dictionary_row = {"Title":newsTitle,"Citation":newsCitation,"MiniSummary":newsMiniSummary,"IssueDate":newsDate,"PubDate":newsPubDate,"Sources":newsSources,"URLs":newsURLString,"PA":newsPA.text,"DatesDifferent":DatesDifferent, "Link":newsLink}
                df = df.append(dictionary_row, ignore_index=True)        
                log(str(newsCitation))    
        else:
            log('NOT ADDED - Citation found in last week archive: ' + str(newsCitation))
    df['TopicName'] = ''
    columnsTitles = ['TopicName', 'Title', 'MiniSummary', 'IssueDate', 'PubDate', 'Sources', 'URLs', 'Citation', 'PA', 'Link', 'DatesDifferent']
    df = df.reindex(columns=columnsTitles) #reorder columns
    df.to_csv(logDir + 'all-pas-news-list.csv', index=False)
    df.to_csv(archive_filename, index=False)
    return df

   
def FindMostRecentFile(directory, pattern):
    try:
        filelist = os.path.join(directory, pattern) #builds list of file in a directory based on a pattern
        filelist = sorted(glob.iglob(filelist), key=os.path.getmtime, reverse=True) #sort by date modified with the most recent at the top
        return filelist[0]
    except: return 'na'

def log(message):
    l = open(JCSLogFile,'a')
    l.write(message + '\n')
    l.close()
    print(message)

#state = 'livedev'
state = 'live'

if state == 'livedev':
    reportDir = '\\\\atlas\\Knowhow\\AutomatedContentReports\\NewsReport\\'  
    logDir = "\\\\atlas\\lexispsl\\Highlights\\dev\\logs\\"
    outputDir = '\\\\atlas\\lexispsl\\Highlights\\dev\\Practice Areas\\'
    TodaysDate = datetime.datetime.today()
    #TodaysDate = datetime.date(2020, 5, 6)

if state == 'live':
    reportDir = '\\\\atlas\\Knowhow\\AutomatedContentReports\\NewsReport\\'
    logDir = "\\\\atlas\\lexispsl\\Highlights\\logs\\"    
    outputDir = '\\\\atlas\\lexispsl\\Highlights\\Practice Areas\\'
    TodaysDate = datetime.datetime.today()
    #TodaysDate = datetime.date(2020, 2, 27)
    
highlightsArchiveDir = '\\\\atlas\\Knowhow\\HighlightsArchive\\'

search = 'https://www.lexisnexis.com/uk/lexispsl/tax/search?pa=arbitration%2Cbankingandfinance%2Ccommercial%2Ccompetition%2Cconstruction%2Ccorporate%2Ccorporatecrime%2Cdisputeresolution%2Cemployment%2Cenergy%2Cenvironment%2Cfamily%2Cfinancialservices%2Cimmigration%2Cinformationlaw%2Cinhouseadvisor%2Cinsuranceandreinsurance%2Cip%2Clifesciences%2Clocalgovernment%2Cpensions%2Cpersonalinjury%2Cplanning%2Cpracticecompliance%2Cprivateclient%2Cproperty%2Cpropertydisputes%2Cpubliclaw%2Crestructuringandinsolvency%2Criskandcompliance%2Ctax%2Ctmt%2Cwillsandprobate&submitOnce=true&wa_origin=paHomePage&wordwheelFaces=daAjax%2Fwordwheel.faces&query='

AllPAs = ['Arbitration', 'Banking and Finance', 'Commercial', 'Competition', 'Construction', 'Corporate', 'Corporate Crime', 'Dispute Resolution', 'Employment', 'Energy', 'Environment', 'Family', 'Financial Services', 'Immigration', 'Information Law', 'In-House Advisor', 'Insurance', 'IP', 'Life Sciences and Pharmaceuticals', 'Local Government', 'Pensions', 'Personal Injury', 'Planning', 'Practice Compliance', 'Practice Management', 'Private Client', 'Property', 'Property Disputes', 'Public Law', 'Restructuring and Insolvency', 'Risk and Compliance', 'Share Schemes', 'Tax', 'TMT', 'Wills and Probate']    
#AllPAs = ['Arbitration', 'Banking &amp; Finance', 'Commercial', 'Competition', 'Construction', 'Corporate', 'Corporate Crime', 'Dispute Resolution', 'Employment', 'Energy', 'Environment', 'Family', 'Financial Services', 'Immigration', 'Information Law', 'In-House Advisor', 'Insurance', 'IP', 'Life Sciences', 'Local Government', 'Pensions', 'Personal Injury', 'Planning', 'Practice Compliance', 'Practice Management', 'Private Client', 'Property', 'Property Disputes', 'Public Law', 'Restructuring &amp; Insolvency', 'Risk &amp; Compliance', 'Share Incentives', 'Tax', 'TMT', 'Wills &amp; Probate']    
MonthlyPAs = ['Competition', 'Family', 'Immigration', 'Insurance', 'Practice Compliance', 'Restructuring and Insolvency', 'Risk and Compliance']    
NSMAP = {'lnci': "http://www.lexisnexis.com/namespace/common/lnci", 'core': 'http://www.lexisnexis.com/namespace/sslrp/core', 'fn': 'http://www.lexisnexis.com/namespace/sslrp/fn', 'header': 'http://www.lexisnexis.com/namespace/uk/header', 'kh': 'http://www.lexisnexis.com/namespace/uk/kh', 'lnb': 'http://www.lexisnexis.com/namespace/uk/lnb', 'lnci': 'http://www.lexisnexis.com/namespace/common/lnci', 'tr': 'http://www.lexisnexis.com/namespace/sslrp/tr'}#, 'atict': 'http://www.arbortext.com/namespace/atict'}

JCSLogFile = logDir + 'JCSlog-newscsvgen.txt'
#JCSLogFile = logDir + time.strftime("JCSlog-newscsvgen-%d%m%Y-%H%M%S.txt")

#with open(JCSLogFile, 'w', newline='', encoding='utf-8') as myfile:

l = open(JCSLogFile,'w')
logdate =  str(time.strftime("%d/%m/%Y %H:%M:%S"))
l.write("Start "+logdate+"\n")
l.close()


print('Writing log here: ' + JCSLogFile)

highlightDate = str(time.strftime("%#d %B %Y %p")) #the hash character turns off the leading zero in the day   
day_and_ampm = str(time.strftime("%A-%p")) #the A gives the week day, the p gives am or pm
archive_filename = logDir + 'all-pas-news-list-' + day_and_ampm + '.csv'
thursday_archive_filename = logDir + 'all-pas-news-list-Thursday-AM.csv'

#main script
log("Today's date is: " + highlightDate)
newsListFilepath = FindMostRecentFile(reportDir, '*.xml')
#load last thursday's csv file to check against when building this run's csv
try: df_archive = pd.read_csv(thursday_archive_filename)
except: 
    df_archive = pd.DataFrame(columns = ['Citation']) #set up blank dataframe if no archive exists, helps further down the line
    log('No archive csv log found: ' + archive_filename)

extract_from_news_feed(df_archive)
df = pd.read_csv(logDir + 'all-pas-news-list.csv')
#wait = input("PAUSED...when ready press enter")

log("\nNews CSV guide generation for weekly highlights...\n")

for PA in AllPAs:
    PrevHighlightsFilepath = FindLastWeekHighLightDoc(highlightsArchiveDir, PA)
    #if PA not in MonthlyPAs: 
    csv_generation(PA, highlightDate, 'weekly', df, outputDir, PrevHighlightsFilepath)
    xls_add_format(PA)

print('Finished, access the log here: ' + JCSLogFile)
log('Finished')


